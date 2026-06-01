from __future__ import annotations
import io
from collections import defaultdict
from sqlalchemy import select
from ..extensions import db


def build_allocation_tsv() -> bytes:
    """TSVを返す

    計算手順:
    """
    headers = ['地区', '課コード', '工程コード', '原価センタ', '編成', '固定',
               '人員計', '人員比', '工程配賦', '端数調整', '工程配賦計']
    rows: list[list] = [headers]

    return _to_tsv_bytes(rows)


def build_labor_tsv(unit_price: float) -> bytes:
    """TSVを返す

    振替金額 = ROUND(作業時間 × unit_price, 0)
    """
    headers = ['勘定科目コード', '原価センタ', '負担課', '担当課', '工事名',
               '作業時間', '振替金額', 'WBS', '資産集約番号', '指図', '備考']
    rows: list[list] = [headers]

    return _to_tsv_bytes(rows)


def build_labor_tran_mock_excel(district_codes: list[str]) -> bytes:
    """選択された地区コードを記載したモック労務費トランExcelを返す。

    NOTE: 集計ロジックは未実装です。このファイルはモックです。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError('openpyxl が必要です: pip install openpyxl')

    wb = Workbook()
    ws = wb.active
    ws.title = '労務費トラン'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='4472C4')
    center = Alignment(horizontal='center')

    # NOTE行
    ws['A1'] = '※ このファイルはモックです。集計ロジックは未実装です。'
    ws['A1'].font = Font(bold=True, color='FF0000')
    ws.merge_cells('A1:D1')

    # ヘッダ行
    headers = ['No.', '地区コード', '労務費計（円）', '備考']
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # データ行
    for i, code in enumerate(district_codes, start=1):
        ws.cell(row=i + 2, column=1, value=i)
        ws.cell(row=i + 2, column=2, value=code)
        ws.cell(row=i + 2, column=3, value=0)
        ws.cell(row=i + 2, column=4, value='集計ロジック未実装')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_tsv_bytes(rows: list[list]) -> bytes:
    """行リストをタブ区切り・CRLF・cp932 のバイト列に変換する。"""
    buf = io.StringIO()
    for row in rows:
        buf.write('\t'.join(str(v) for v in row) + '\r\n')
    return buf.getvalue().encode('cp932', errors='replace')
