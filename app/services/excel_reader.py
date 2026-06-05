from __future__ import annotations

import re
import os
import yaml
import openpyxl
from pathlib import Path
from typing import Any
from openpyxl.utils import column_index_from_string
from pydantic import BaseModel, ValidationError
from dataclasses import dataclass, field
from openpyxl import load_workbook
from ..validators import get_model

_YAML_PATH = Path(__file__).parent.parent.parent / 'config' / 'excel_formats.yaml'
_YEAR_MONTH_RE = re.compile(r"(\d{4})[年/\-](\d{1,2})[月]?")
# "B7:I7" → ('B', 7, 'I', 7)
_RANGE_RE = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", re.IGNORECASE)


class ColumnDef(BaseModel):
    name: str
    type: str
    required: bool = False
    description: str | None = None


class ExcelFormatConfig(BaseModel):
    display_name: str
    sheet_name: str
    year_cell: str | None = None
    month_cell: str | None = None
    header_cell_range: str
    columns: list[ColumnDef]


@dataclass
class RowError:
    row_number: int
    data: dict[str, Any]
    error: str

@dataclass
class ExcelReadResult:
    file_type: str
    year: int | None
    month: int | None
    rows: list[BaseModel]
    errors: list[RowError] = field(default_factory=list)

def load_formats(yaml_path: str | Path) -> dict[str, ExcelFormatConfig]:
    with open(yaml_path, encoding="utf-8") as f:
        raw = yaml.safe_load(f)
    return {key: ExcelFormatConfig.model_validate(val) for key, val in raw.items()}

def get_format_config(file_type: str) -> ExcelFormatConfig:
    formats = load_formats(_YAML_PATH)
    if file_type not in formats:
        raise ValueError(f"未定義のファイル種別です: {file_type}")
    return formats[file_type]

class ExcelReader:
    def __init__(self, formats_yaml: str | Path = "excel_formats.yaml") -> None:
        self.formats: dict[str, ExcelFormatConfig] = load_formats(formats_yaml)

    def read(self, excel_path: str | Path, file_type: str) -> ExcelReadResult:
        config = self.formats[file_type]
        row_model = get_model(file_type)

        wb = load_workbook(excel_path, data_only=True)
        ws = wb[config.sheet_name]

        year, month = self._extract_year_month(ws, config)
        headers, data_start_row = self._read_headers(ws, config.header_cell_range)
        
        rows: list[BaseModel] = []
        errors: list[RowError] = []
        consecutive_empty = 0

        for row_idx in range(data_start_row, ws.max_row + 1):
            row_data = {
                headers[col_idx]: ws.cell(row=row_idx, column=col_idx).value
                for col_idx in headers
            }
            print(f"Row {row_idx}: {row_data}")  # デバッグ用出力

            # 全列Noneの空行が2行続いたら終了
            if all(v is None for v in row_data.values()):
                consecutive_empty += 1
                if consecutive_empty >= 2:
                    break
                continue
            consecutive_empty = 0

            try:
                rows.append(row_model.model_validate(row_data))
            except ValidationError as exc:
                print(exc)
                errors.append(
                    RowError(
                        row_number=row_idx,
                        data=row_data,
                        error=exc.json(indent=None),
                    )
                )

        return ExcelReadResult(
            file_type=file_type,
            year=year,
            month=month,
            rows=rows,
            errors=errors,
        )

    # ------------------------------------------------------------------

    def _parse_year_month(self, value: object) -> tuple[int | None, int | None]:
        """セル値から (year, month) を抽出する。"""
        if value is None:
            return None, None
        text = str(value)
        m = _YEAR_MONTH_RE.search(text)
        if m:
            return int(m.group(1)), int(m.group(2))
        return None, None
    

    def _extract_year_month(
        self, ws: Any, config: ExcelFormatConfig
    ) -> tuple[int | None, int | None]:
        year: int | None = None
        month: int | None = None

        if config.year_cell and config.month_cell and config.year_cell == config.month_cell:
            # 同一セル（例: ouen の C1）
            val = ws[config.year_cell].value
            year, month = self._parse_year_month(val)
        else:
            if config.year_cell:
                year, _ = self._parse_year_month(ws[config.year_cell].value)
            if config.month_cell:
                _, month = self._parse_year_month(ws[config.month_cell].value)

        return year, month

    def _read_headers(
        self, ws: Any, header_cell_range: str
    ) -> tuple[dict[int, str], int]:
        """
        Returns:
            headers: {col_index: header_name}
            data_start_row: ヘッダー行の次の行番号
        """
        m = _RANGE_RE.match(header_cell_range.strip())
        if not m:
            raise ValueError(f"header_cell_range の形式が不正です: {header_cell_range!r}")

        start_col = column_index_from_string(m.group(1))
        header_row = int(m.group(2))
        end_col = column_index_from_string(m.group(3))

        headers: dict[int, str] = {}
        seen: dict[str, int] = {}
        for col_idx in range(start_col, end_col + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if val is not None:
                name = str(val).replace('\u3000', '').replace(' ', '').strip()
                if name in seen:
                    seen[name] += 1
                    name = f"{name}_{seen[name]}"
                else:
                    seen[name] = 1
                headers[col_idx] = name

        return headers, header_row + 1


