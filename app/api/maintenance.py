import io

import openpyxl
from flask import render_template, request, flash, redirect, url_for, send_file, session
from flask_login import login_required
from sqlalchemy import select

from ..views import main_bp
from ..extensions import db
from ..models.section import SectionMaster
from ..models.department import DepartmentMaster
from ..services.master_importer import (
    read_and_validate_section,
    read_and_validate_department,
    save_pending,
    load_pending,
    delete_pending,
)
from ..models.district import DistrictMaster
from ..models.account import AccountMaster
from ..models.cost_center import CostCenterMaster
from ..models.salary import SalaryData
from ..models.allocation import AllocationData
from ..models.labor_transfer import LaborTransferData
from ..models.ouen import OuenData


# ---- Section Master ----

@main_bp.route('/maintenance/section/create', methods=['POST'])
@login_required
def section_create():
    section = SectionMaster(
        section_code=request.form['section_code'].strip(),
        section_name=request.form['section_name'].strip(),
        district_code=request.form['district_code'].strip(),
        cost_center_code=request.form['cost_center_code'].strip(),
    )
    db.session.add(section)
    try:
        db.session.commit()
        flash('課マスタを追加しました。', 'success')
    except Exception:
        db.session.rollback()
        flash('課コードが重複しています。', 'danger')
    return redirect(url_for('main.section_list'))


@main_bp.route('/maintenance/section/update/<section_code>', methods=['POST'])
@login_required
def section_update(section_code: str):
    section = db.get_or_404(SectionMaster, section_code)
    section.section_name = request.form['section_name'].strip()
    section.district_code = request.form['district_code'].strip()
    section.cost_center_code = request.form['cost_center_code'].strip()
    db.session.commit()
    flash('課マスタを更新しました。', 'success')
    return redirect(url_for('main.section_list'))


@main_bp.route('/maintenance/section/delete/<section_code>', methods=['POST'])
@login_required
def section_delete(section_code: str):
    section = db.get_or_404(SectionMaster, section_code)
    db.session.delete(section)
    db.session.commit()
    flash('課マスタを削除しました。', 'success')
    return redirect(url_for('main.section_list'))


# ---- Department Master ----

@main_bp.route('/maintenance/department/create', methods=['POST'])
@login_required
def department_create():
    dept = DepartmentMaster(
        department_code=request.form['department_code'].strip(),
        department_name=request.form['department_name'].strip(),
        district_code=request.form['district_code'].strip(),
        section_code=request.form['section_code'].strip(),
        agg_section_code=request.form['agg_section_code'].strip(),
        kbn_code=request.form['kbn_code'].strip(),
        account_code=request.form['account_code'].strip(),
        cost_center_code=request.form['cost_center_code'].strip(),
    )
    db.session.add(dept)
    try:
        db.session.commit()
        flash('変換マスタを追加しました。', 'success')
    except Exception:
        db.session.rollback()
        flash('人事所属コードが重複しています。', 'danger')
    return redirect(url_for('main.department_list'))


@main_bp.route('/maintenance/department/update/<department_code>', methods=['POST'])
@login_required
def department_update(department_code: str):
    dept = db.get_or_404(DepartmentMaster, department_code)
    dept.department_name = request.form['department_name'].strip()
    dept.district_code = request.form['district_code'].strip()
    dept.section_code = request.form['section_code'].strip()
    dept.agg_section_code = request.form['agg_section_code'].strip()
    dept.kbn_code = request.form['kbn_code'].strip()
    dept.account_code = request.form['account_code'].strip()
    dept.cost_center_code = request.form['cost_center_code'].strip()
    db.session.commit()
    flash('変換マスタを更新しました。', 'success')
    return redirect(url_for('main.department_list'))


@main_bp.route('/maintenance/department/delete/<department_code>', methods=['POST'])
@login_required
def department_delete(department_code: str):
    dept = db.get_or_404(DepartmentMaster, department_code)
    db.session.delete(dept)
    db.session.commit()
    flash('変換マスタを削除しました。', 'success')
    return redirect(url_for('main.department_list'))


# ---- Section Master Import ----

_SECTION_IMPORT_SESSION_KEY = 'section_import_uuid'
_DEPARTMENT_IMPORT_SESSION_KEY = 'department_import_uuid'


@main_bp.route('/maintenance/section/import', methods=['POST'])
@login_required
def section_import():
    file = request.files.get('import_file')
    if not file or not file.filename:
        flash('ファイルを選択してください。', 'danger')
        return redirect(url_for('main.section_list'))
    if not file.filename.lower().endswith('.xlsx'):
        flash('xlsx ファイルを選択してください。', 'danger')
        return redirect(url_for('main.section_list'))

    current_count = db.session.scalar(select(db.func.count()).select_from(SectionMaster)) or 0
    result = read_and_validate_section(file, current_count)

    if not result.ok:
        for msg in result.errors:
            flash(msg, 'danger')
        return redirect(url_for('main.section_list'))

    if result.has_warnings:
        # Discard any previous pending import for this user
        old_token = session.pop(_SECTION_IMPORT_SESSION_KEY, None)
        if old_token:
            delete_pending(old_token, 'section')
        token = save_pending(result.rows, 'section')
        session[_SECTION_IMPORT_SESSION_KEY] = token
        for msg in result.warnings:
            flash(msg, 'warning')
        return redirect(url_for('main.section_list', confirm='section'))

    # No warnings — replace immediately
    _replace_section_master(result.rows)
    flash(f'課マスタを取り込みました（{len(result.rows)} 件）。', 'success')
    return redirect(url_for('main.section_list'))


@main_bp.route('/maintenance/section/import/confirm', methods=['POST'])
@login_required
def section_import_confirm():
    token_form = request.form.get('import_uuid', '')
    token_session = session.get(_SECTION_IMPORT_SESSION_KEY, '')
    if not token_form or token_form != token_session:
        flash('取り込みセッションが無効です。再度ファイルを選択してください。', 'danger')
        return redirect(url_for('main.section_list'))

    rows = load_pending(token_form, 'section')
    if rows is None:
        flash('取り込みデータが見つかりません。再度ファイルを選択してください。', 'danger')
        return redirect(url_for('main.section_list'))

    _replace_section_master(rows)
    delete_pending(token_form, 'section')
    session.pop(_SECTION_IMPORT_SESSION_KEY, None)
    flash(f'課マスタを取り込みました（{len(rows)} 件）。', 'success')
    return redirect(url_for('main.section_list'))


def _replace_section_master(rows: list[dict]) -> None:
    db.session.execute(db.delete(SectionMaster))
    db.session.bulk_save_objects([
        SectionMaster(
            section_code=r['section_code'],
            section_name=r['section_name'],
            district_code=r['district_code'],
            cost_center_code=r['cost_center_code'],
        )
        for r in rows
    ])
    db.session.commit()


# ---- Department Master Import ----


@main_bp.route('/maintenance/department/import', methods=['POST'])
@login_required
def department_import():
    file = request.files.get('import_file')
    if not file or not file.filename:
        flash('ファイルを選択してください。', 'danger')
        return redirect(url_for('main.department_list'))
    if not file.filename.lower().endswith('.xlsx'):
        flash('xlsx ファイルを選択してください。', 'danger')
        return redirect(url_for('main.department_list'))

    current_count = db.session.scalar(select(db.func.count()).select_from(DepartmentMaster)) or 0
    result = read_and_validate_department(file, current_count)

    if not result.ok:
        for msg in result.errors:
            flash(msg, 'danger')
        return redirect(url_for('main.department_list'))

    if result.has_warnings:
        old_token = session.pop(_DEPARTMENT_IMPORT_SESSION_KEY, None)
        if old_token:
            delete_pending(old_token, 'department')
        token = save_pending(result.rows, 'department')
        session[_DEPARTMENT_IMPORT_SESSION_KEY] = token
        for msg in result.warnings:
            flash(msg, 'warning')
        return redirect(url_for('main.department_list', confirm='department'))

    _replace_department_master(result.rows)
    flash(f'変換マスタを取り込みました（{len(result.rows)} 件）。', 'success')
    return redirect(url_for('main.department_list'))


@main_bp.route('/maintenance/department/import/confirm', methods=['POST'])
@login_required
def department_import_confirm():
    token_form = request.form.get('import_uuid', '')
    token_session = session.get(_DEPARTMENT_IMPORT_SESSION_KEY, '')
    if not token_form or token_form != token_session:
        flash('取り込みセッションが無効です。再度ファイルを選択してください。', 'danger')
        return redirect(url_for('main.department_list'))

    rows = load_pending(token_form, 'department')
    if rows is None:
        flash('取り込みデータが見つかりません。再度ファイルを選択してください。', 'danger')
        return redirect(url_for('main.department_list'))

    _replace_department_master(rows)
    delete_pending(token_form, 'department')
    session.pop(_DEPARTMENT_IMPORT_SESSION_KEY, None)
    flash(f'変換マスタを取り込みました（{len(rows)} 件）。', 'success')
    return redirect(url_for('main.department_list'))


def _replace_department_master(rows: list[dict]) -> None:
    db.session.execute(db.delete(DepartmentMaster))
    db.session.bulk_save_objects([
        DepartmentMaster(
            department_code=r['department_code'],
            department_name=r['department_name'],
            district_code=r['district_code'],
            section_code=r['section_code'],
            agg_section_code=r['agg_section_code'],
            kbn_code=r['kbn_code'],
            account_code=r['account_code'],
            cost_center_code=r['cost_center_code'],
        )
        for r in rows
    ])
    db.session.commit()


# ---- Data Output ----

_ALLOWED_TABLES = {
    'salary':         (SalaryData,        ['id', 'batch_id', 'department_code', 'base_salary', 'ability_salary', 'compensation', 'allowance', 'headcount', 'total_salary', 'created_at', 'created_by']),
    'ouen':           (OuenData,          ['id', 'batch_id', 'from_district', 'from_section_code', 'to_district', 'to_section_code', 'departure_date', 'return_date', 'days', 'extended_days', 'note', 'created_at', 'created_by']),
    'allocation':     (AllocationData,    ['id', 'batch_id', 'district_code', 'section_code', 'allocation_ratio', 'created_at', 'created_by']),
    'labor_transfer': (LaborTransferData, ['id', 'batch_id', 'account_code', 'from_section_code', 'to_section_code', 'work_hours', 'created_at', 'created_by']),
    'section':        (SectionMaster,     ['section_code', 'section_name', 'district_code', 'cost_center_code']),
    'department':     (DepartmentMaster,  ['department_code', 'department_name', 'district_code', 'section_code', 'agg_section_code', 'kbn_code', 'account_code', 'cost_center_code']),
    'district':       (DistrictMaster,    ['district_code', 'district_name']),
    'account':        (AccountMaster,     ['account_code', 'account_name']),
    'cost_center':    (CostCenterMaster,  ['cost_center_code', 'cost_center_name']),
}


def _write_sheet(ws, columns: list[str], records) -> None:
    ws.append(columns)
    for rec in records:
        ws.append([getattr(rec, col, '') for col in columns])


@main_bp.route('/maintenance/data-output/download/<table_name>')
@login_required
def data_output_download(table_name: str):
    if table_name not in _ALLOWED_TABLES:
        flash('無効なテーブルです。', 'danger')
        return redirect(url_for('main.data_output'))

    model_cls, columns = _ALLOWED_TABLES[table_name]
    records = db.session.scalars(select(model_cls)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_name
    _write_sheet(ws, columns, records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{table_name}.xlsx',
    )


@main_bp.route('/maintenance/data-output/download-all')
@login_required
def data_output_download_all():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for table_name, (model_cls, columns) in _ALLOWED_TABLES.items():
        records = db.session.scalars(select(model_cls)).all()
        ws = wb.create_sheet(title=table_name)
        _write_sheet(ws, columns, records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='all_tables.xlsx',
    )
