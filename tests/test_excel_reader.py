"""Tests for the Excel reader service."""
import os
import tempfile
import pytest
import openpyxl

from app.services.excel_reader import read_excel, get_format_config


def _make_salary_xlsx() -> str:
    """Create a minimal salary Excel file and return its path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '給与データ'
    # Header row (row 1) — matches actual Excel column layout
    ws.append([
        '地区', '課コード', '地区+課コード', '集約課コード', '地区+集約課コード',
        '原価センタ', '区分', '勘定科目', '行ラベル', '所属名',
        '合計/(明細1)本給', '合計/(明細1)能力給', '合計/(明細1)職務・役割給', '合計/(明細1)役割業績給',
    ])
    # Data rows
    ws.append(['東京', '5020', '東京+5020', '5020', '東京+5020', 'CC001', 'A', '4100', '00000100', '製造一課', 300000, 50000, 20000, 30000])
    ws.append(['大阪', '5030', '大阪+5030', '5030', '大阪+5030', 'CC002', 'B', '4200', '00000200', '製造二課', 250000, 40000, 15000, 25000])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def _make_allocation_xlsx() -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '工程配賦'
    # header_row=1
    ws.append(['事業部', '地区', '課コード', '原価区分', '工程', '日数', '工程名', '編成', '固定'])
    ws.append(['50', '10', '5020', 'A', '200', 20.0, '工程A', 0.2, 0.0])
    ws.append(['50', '10', '5020', 'A', '260', 20.0, None, 0.1, 0.0])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


class TestGetFormatConfig:
    def test_salary_config_loaded(self):
        cfg = get_format_config('salary')
        assert cfg['sheet_name'] == '給与データ'
        assert len(cfg['columns']) == 40

    def test_unknown_type_raises(self):
        with pytest.raises(ValueError, match='未定義'):
            get_format_config('unknown_type')


class TestReadExcel:
    def test_read_salary(self):
        path = _make_salary_xlsx()
        try:
            rows = read_excel(path, 'salary')
            assert len(rows) == 2
            assert rows[0]['行ラベル'] == '00000100'
            assert rows[0]['合計/(明細1)本給'] == 300000
        finally:
            os.unlink(path)

    def test_read_allocation(self):
        path = _make_allocation_xlsx()
        try:
            rows = read_excel(path, 'allocation')
            assert len(rows) == 2
            assert rows[0]['課コード'] == '5020'
            assert rows[0]['編成'] == pytest.approx(0.2)
        finally:
            os.unlink(path)

    def test_missing_sheet_raises(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'WrongSheet'
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(tmp.name)
        tmp.close()
        try:
            with pytest.raises(ValueError, match='シート'):
                read_excel(tmp.name, 'salary')
        finally:
            os.unlink(tmp.name)
