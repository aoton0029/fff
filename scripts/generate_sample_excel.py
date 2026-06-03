"""
config/excel_formats.yaml の定義に基づいてアップロード用サンプルExcelを生成する。
出力先: instance/upload/
"""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = ROOT / "config" / "excel_formats.yaml"
OUTPUT_DIR = ROOT / "instance" / "upload"

SAMPLE_DATA: dict[str, list[dict]] = {
    "salary": [
        {
            "行ラベル": "12345678",
            "所属名": "製造第一課",
            "合計/(明細1)本給": 250000,
            "合計/(明細1)能力給": 30000,
            "合計/(明細1)職務・役割給": 20000,
            "合計/(明細1)役割業績給": 15000,
            "合計/(明細1)報酬": 0,
            "合計/(明細1)本俸": 0,
            "合計/(明細1)基本賃金": 280000,
            "合計/(明細1)国内給与": 0,
            "合計/(明細1)駐在員給与調整": 0,
            "合計/(明細1)家族手当": 10000,
            "合計/(明細1)資格手当": 5000,
            "合計/(明細1)交替勤務加算給": 8000,
            "合計/(明細1)振替勤務就業手当": 0,
            "合計/(明細1)早出残業手当": 12000,
            "合計/(明細1)休日出勤手当": 0,
            "合計/(明細1)特定休日出勤手当": 0,
            "合計/(明細1)交替勤務手当": 6000,
            "合計/(明細1)守衛手当": 0,
            "合計/(明細1)深夜業手当": 4000,
            "合計/(明細1)単身赴任手当": 0,
            "合計/(明細1)一時帰省旅費": 0,
            "合計/(明細1)夜間呼出手当": 0,
            "合計/(明細1)出向補償時間": 0,
            "合計/(明細1)その他手当": 0,
            "合計/(明細1)前月育介金額": 0,
            "合計/(明細1)公的資格手当": 3000,
            "合計/(明細1)賃金控除": 0,
            "合計/(明細2)職場離脱(会計)": 0,
            "合計/(明細2)人件費人数(会計)": 1,
        },
        {
            "行ラベル": "87654321",
            "所属名": "製造第二課",
            "合計/(明細1)本給": 220000,
            "合計/(明細1)能力給": 25000,
            "合計/(明細1)職務・役割給": 18000,
            "合計/(明細1)役割業績給": 12000,
            "合計/(明細1)報酬": 0,
            "合計/(明細1)本俸": 0,
            "合計/(明細1)基本賃金": 245000,
            "合計/(明細1)国内給与": 0,
            "合計/(明細1)駐在員給与調整": 0,
            "合計/(明細1)家族手当": 15000,
            "合計/(明細1)資格手当": 0,
            "合計/(明細1)交替勤務加算給": 0,
            "合計/(明細1)振替勤務就業手当": 0,
            "合計/(明細1)早出残業手当": 9000,
            "合計/(明細1)休日出勤手当": 5000,
            "合計/(明細1)特定休日出勤手当": 0,
            "合計/(明細1)交替勤務手当": 0,
            "合計/(明細1)守衛手当": 0,
            "合計/(明細1)深夜業手当": 0,
            "合計/(明細1)単身赴任手当": 20000,
            "合計/(明細1)一時帰省旅費": 0,
            "合計/(明細1)夜間呼出手当": 0,
            "合計/(明細1)出向補償時間": 0,
            "合計/(明細1)その他手当": 2000,
            "合計/(明細1)前月育介金額": 0,
            "合計/(明細1)公的資格手当": 0,
            "合計/(明細1)賃金控除": 0,
            "合計/(明細2)職場離脱(会計)": 0,
            "合計/(明細2)人件費人数(会計)": 2,
        },
    ],
    "allocation": [
        {
            "事業部": "製造",
            "地区": "関東",
            "課コード": "1234",
            "原価区分": "直接",
            "工程": "A001",
            "日数": 20.0,
            "工程名": "組立工程",
            "編成": 5.0,
            "固定": 2.0,
        },
        {
            "事業部": "製造",
            "地区": "関東",
            "課コード": "1234",
            "原価区分": "間接",
            "工程": "B002",
            "日数": 20.0,
            "工程名": "検査工程",
            "編成": 3.0,
            "固定": 1.0,
        },
        {
            "事業部": "製造",
            "地区": "関西",
            "課コード": "5678",
            "原価区分": "直接",
            "工程": "A003",
            "日数": 22.0,
            "工程名": "塗装工程",
            "編成": 4.0,
            "固定": 1.5,
        },
    ],
    "labor_transfer": [
        {
            "勘定科目コード": "51010001",
            "原価センタ": "CC001",
            "負担課": "1234",
            "担当課": "5678",
            "工事名": "第一工場設備改修",
            "作業時間": 40.0,
            "WBS": "WBS-001",
            "資産集約番号": "",
            "指図": "",
            "備考": "5月分",
        },
        {
            "勘定科目コード": "51020002",
            "原価センタ": "CC002",
            "負担課": "5678",
            "担当課": "9012",
            "工事名": "第二工場電気工事",
            "作業時間": 24.5,
            "WBS": "",
            "資産集約番号": "A-2024-001",
            "指図": "OR-001",
            "備考": "",
        },
    ],
    "ouen": [
        {
            "課コード": "1234",
            "課名": "製造第一課",
            "課コード_2": "5678",
            "課名_2": "製造第二課",
            "人数": 2,
            "出課日": date(2026, 5, 1),
            "帰課日": date(2026, 5, 10),
            "日数": 10,
            "延日数": 20,
            "氏名": "山田 太郎",
        },
        {
            "課コード": "5678",
            "課名": "製造第二課",
            "課コード_2": "9012",
            "課名_2": "製造第三課",
            "人数": 1,
            "出課日": date(2026, 5, 15),
            "帰課日": date(2026, 5, 20),
            "日数": 6,
            "延日数": 6,
            "氏名": "鈴木 花子",
        },
    ],
}

MONTH_CELL_VALUE = "2026/05"
OUEN_YEAR_MONTH_VALUE = "2026年5月"


def col_letter_to_index(letter: str) -> int:
    """列記号（A, B, ..., AM）を1始まりのインデックスに変換する。"""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def parse_use_cols(use_cols: str) -> tuple[int, int]:
    """'B:I' のような文字列を (start_col_index, end_col_index) に変換する。"""
    parts = use_cols.split(":")
    start = col_letter_to_index(re.sub(r"\d", "", parts[0]))
    end = col_letter_to_index(re.sub(r"\d", "", parts[1]))
    return start, end


def write_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    fmt: dict,
    rows: list[dict],
    *,
    extra_cells: dict[str, object] | None = None,
) -> None:
    """ワークシートにヘッダーとサンプルデータを書き込む。"""
    if extra_cells:
        for cell_ref, value in extra_cells.items():
            ws[cell_ref] = value

    start_col, _ = parse_use_cols(fmt["use_cols"])
    header_row: int = fmt["header_row"]
    col_names = [c["name"] for c in fmt["columns"]]

    # ヘッダー行
    for i, name in enumerate(col_names):
        ws.cell(row=header_row, column=start_col + i, value=name)

    # データ行（ouenは課コードが重複するため values() の順序で書き込む）
    for row_offset, row_data in enumerate(rows, start=1):
        values = list(row_data.values())
        for i, value in enumerate(values):
            ws.cell(row=header_row + row_offset, column=start_col + i, value=value)


def generate(fmt_key: str, fmt: dict, rows: list[dict], filename: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = fmt["sheet_name"]

    extra_cells: dict[str, object] = {}
    if fmt.get("month_cell"):
        extra_cells[fmt["month_cell"]] = MONTH_CELL_VALUE
    if fmt.get("year_cell") and fmt["year_cell"] != fmt.get("month_cell"):
        extra_cells[fmt["year_cell"]] = OUEN_YEAR_MONTH_VALUE
    if fmt_key == "ouen":
        extra_cells["C1"] = OUEN_YEAR_MONTH_VALUE

    write_sheet(ws, fmt, rows, extra_cells=extra_cells or None)

    out_path = OUTPUT_DIR / filename
    wb.save(out_path)
    print(f"  作成: {out_path}")


def main() -> None:
    with open(CONFIG_PATH, encoding="utf-8") as f:
        config: dict = yaml.safe_load(f)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    mapping = {
        "salary": "sample_salary.xlsx",
        "allocation": "sample_allocation.xlsx",
        "labor_transfer": "sample_labor_transfer.xlsx",
        "ouen": "sample_ouen.xlsx",
    }

    for key, filename in mapping.items():
        fmt = config[key]
        rows = SAMPLE_DATA[key]
        generate(key, fmt, rows, filename)

    print("完了")


if __name__ == "__main__":
    main()
